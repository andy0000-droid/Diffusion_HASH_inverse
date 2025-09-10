"""
File I/O Utilities
"""
from __future__ import annotations
from dataclasses import dataclass
from pathlib import Path
from datetime import datetime
import os
import re

import pandas as pd
from openpyxl import load_workbook
from pandas import ExcelWriter

from diffusion_hash_inv.utils.project_root import add_src_to_path, add_root_to_path
add_src_to_path()
ROOT_DIR = add_root_to_path()

# 고정 헤더 길이
TS_LEN = 32
BITLEN_LEN = 8 # 64 bits
DIFFTIME_LEN = 8 # 64 bits
PAD_LEN = 16 - BITLEN_LEN - DIFFTIME_LEN  # 8 bytes padding to make total 48 bytes
HEADER_LEN = TS_LEN + BITLEN_LEN + PAD_LEN  # 48 bytes

@dataclass
class Header:
    """
    Represents the header of a binary file.
    """
    timestamp_utf8: bytes  # 32B
    time_diff: bytes   # float
    bit_length: int        # uint64

    @classmethod
    def now(cls, bit_length: int, timestamp: bytes = None, time_diff: bytes = None) -> Header:
        """
        Create a new header with the current timestamp and the given bit length.
        """
        assert timestamp is not None, "Timestamp must be provided"
        assert time_diff is not None, "Time difference must be provided"
        return cls(timestamp, time_diff, bit_length)

    def to_bytes(self) -> bytes:
        """
        Convert the header to bytes.
        """
        return self.timestamp_utf8 + self.time_diff + FileIO.encode_bit_length(self.bit_length)

    @classmethod
    def from_file(cls, path: Path) -> Header:
        """
        Create a Header instance from a binary file.
        """
        with open(path, "rb") as f:
            raw = f.read(HEADER_LEN)
        if len(raw) != HEADER_LEN:
            raise ValueError("file too small to contain header")
        ts = raw[:TS_LEN]
        difftime = raw[TS_LEN:TS_LEN + DIFFTIME_LEN]
        bl = FileIO.decode_bit_length(
            raw[TS_LEN + DIFFTIME_LEN:TS_LEN + DIFFTIME_LEN + BITLEN_LEN])
        return cls(ts, difftime, bl)

class FileIO:
    """
    File I/O Utilities
    """
    #pylint: disable=too-many-arguments, too-many-positional-arguments
    def __init__(self, init_flag, clear_flag, verbose_flag, length = 0, start_time: bytes = None):
        self.data_dir = ROOT_DIR / "data"
        self.out_dir = ROOT_DIR / "output"
        if init_flag:
            self.start = self.encode_timestamp()
            self.length = length
        else:
            assert start_time is not None, "start_time must be provided if init_flag is False"
            self.start = start_time
            self.length = None

        if clear_flag:
            print("Clearing generated files...")
            self.file_clean(clear_flag=clear_flag, verbose_flag=verbose_flag)

        self.out_flag = False
        self.json_flag = False

    #pylint: disable=broad-exception-caught, too-many-branches, too-many-nested-blocks
    def file_clean(self, clear_flag = False, verbose_flag = True):
        """
        Clean up the generated files.
        """
        if not clear_flag:
            return

        targets = [self.data_dir, self.out_dir]
        for root_dir in targets:
            if not root_dir:
                continue
            if not os.path.isdir(root_dir):
                if verbose_flag:
                    print(f"[SKIP] 디렉터리가 아님: {root_dir}")
                continue

            # 하위부터 올라오며 파일/디렉터리 삭제
            for cur, dirs, files in os.walk(root_dir, topdown=False, followlinks=False):
                # 파일 삭제
                for name in files:
                    p = os.path.join(cur, name)
                    try:
                        os.unlink(p)  # 파일/하드링크/심볼릭링크 모두 처리
                        if verbose_flag:
                            print(f"Remove file: {p}")
                    except Exception as e:
                        print(f"[SKIP] {p}: {e}")

                # 디렉터리(or 디렉터리 링크) 삭제
                for name in dirs:
                    p = os.path.join(cur, name)
                    try:
                        if os.path.islink(p):
                            os.unlink(p)      # 디렉터리 링크는 unlink
                            if verbose_flag:
                                print(f"Remove symlink dir: {p}")
                        else:
                            if verbose_flag:
                                print(f"Remove directory: {p}")
                            os.rmdir(p)       # 하위가 이미 지워져서 비어 있음
                    except Exception as e:
                        print(f"[SKIP] {p}: {e}")
    #pylint: enable=broad-exception-caught, too-many-branches, too-many-nested-blocks

    @staticmethod
    def _pad16(f):
        pos = f.tell()
        if pos % 16 != 0:
            f.write(b'\x00' * (16 - (pos % 16)))

    @staticmethod
    def encode_timestamp() -> bytes:
        """
        Encode the current timestamp as bytes.
        """
        dt = datetime.now().astimezone()
        tz = dt.strftime("%z")  # +0900
        tz = f"{tz[:3]}:{tz[3:]}"  # +09:00
        s = f"{dt:%Y-%m-%d %H:%M:%S.%f}{tz}"
        if len(s) != TS_LEN:
            raise ValueError(f"timestamp length != {TS_LEN}: {len(s)}")
        return s.encode("utf-8")

    @staticmethod
    def decode_timestamp(b: bytes) -> datetime:
        """
        Decode the timestamp from bytes.
        """
        return datetime.fromisoformat(b.decode('utf-8'))

    @staticmethod
    def _ts_diff(ts1: bytes, ts2: bytes) -> int:
        """
        Calculate the difference between two timestamps.
        ts2 - ts1
        """
        dt1 = FileIO.decode_timestamp(ts1)
        dt2 = FileIO.decode_timestamp(ts2)
        diff_time = dt2 - dt1
        dt = diff_time.days * 86_400_000_000 \
            + diff_time.seconds * 1_000_000 \
            + diff_time.microseconds
        return dt

    @staticmethod
    def encode_bit_length(bit_length: int) -> bytes:
        """
        Encode the bit length as bytes.
        """
        if bit_length < 0:
            raise ValueError("bit_length must be >= 0")
        return bit_length.to_bytes(8, "big", signed=False)

    @staticmethod
    def decode_bit_length(b: bytes) -> int:
        """
        Decode the bit length from bytes.
        """
        return int.from_bytes(b, "big", signed=False)

    def _select_data_dir(self, filename: str) -> Path:
        """
        Decide subdirectory by extension and ensure it exists.
        """
        if filename.endswith(".bin"):
            base = self.data_dir / "binary"
            self.out_flag = False
        elif filename.endswith(".char"):
            base = self.data_dir / "character"
            self.out_flag = False
        elif filename.endswith(".json"):
            base = self.out_dir / "json" / f"{self.length}"
            self.out_flag = True
            self.json_flag = True
        elif filename.endswith(".xlsx"):
            base = self.out_dir / "xlsx" / f"{self.length}"
            self.out_flag = True
            self.json_flag = False
        else:
            raise ValueError("Invalid file extension. Use .bin, .char, .json, or .xlsx")

        base.mkdir(parents=True, exist_ok=True)  # ← 실제 타깃 디렉터리 생성
        return base

    def _sanitize_filename(self, name: str) -> str:
        # 윈도우/범용 안전: 콜론, 슬래시, 역슬래시 등 치환
        return name.replace(":", "-").replace("/", "_").replace("\\", "_")

    #pylint: disable=too-many-statements, broad-exception-caught
    def file_io(self, filename: str):
        """
        Write the given data to a binary file.
        """
        filename = self._sanitize_filename(filename)
        _dir = self._select_data_dir(filename)

        assert _dir is not None, "Invalid file extension. Use .bin, .char, .json, or .xlsx"

        def file_write(*payload, ts: bytes = None):
            """
            Write the data to the file.
            """
            bytes_data, data_len = payload
            time_diff = self._ts_diff(self.start, ts)
            time_diff = time_diff.to_bytes(8, "big", signed=True)

            if isinstance(bytes_data, str):
                bytes_data = bytes.fromhex(bytes_data)  # Convert hex string to bytes

            assert ((data_len + 7) // 8) == len(bytes_data), "Data length mismatch."
            path = _dir / filename
            head = Header.now(data_len, self.start, time_diff)
            try:
                with open(str(path), "ab") as f:
                    self._pad16(f)
                    f.write(head.to_bytes())
                    f.write(bytes_data)
                    self._pad16(f)

            except FileNotFoundError as e:
                print(f"File not found: {e}")

        def json_write(payload_json):
            """
            Write the json to the file
            """

            _path = _dir / filename

            with open(str(_path), "w", encoding="UTF-8", newline="\n") as j:
                j.write(payload_json)

        _illegal = re.compile(r'[\x00-\x08\x0B\x0C\x0E-\x1F]')
        def _sanitize_df(df: pd.DataFrame) -> pd.DataFrame:
            obj = df.select_dtypes(include=['object']).columns
            for c in obj:
                df[c] = df[c].astype(str).map(lambda s: _illegal.sub('', s))
                df[c] = df[c].map(lambda s: "'" + s if s[:1] in ("=","+","-","@") else s)
            return df

        def xlsx_write(payload_xlsx):
            """
            Write the xlsx to the file
            """

            _path = _dir / filename
            _path.parent.mkdir(parents=True, exist_ok=True)
            df = payload_xlsx.copy()
            df = _sanitize_df(df)

            if not _path.exists():
                df.to_excel(str(_path), engine="openpyxl", index=True)
            else:
                wb = load_workbook(str(_path))
                ws = wb["Sheet1"]
                startrow = ws.max_row
                with ExcelWriter(str(_path), engine="openpyxl",
                                mode="a", if_sheet_exists="overlay") as w:
                    df.to_excel(w, sheet_name="Sheet1", startrow=startrow, index=True, header=False)

        def file_read():
            """
            Read the data from the file.
            """
            read_path = self.out_dir / filename
            try:
                with open(read_path, "rb") as f:
                    _ = Header.from_file(read_path)
                    f.seek(HEADER_LEN)
                    _ = f.read()

            except FileNotFoundError as e:
                print(f"File not found: {e}")
    #pylint: enable=too-many-statements
        if self.out_flag:
            if self.json_flag:
                return json_write, file_read
            return xlsx_write, file_read
        return file_write, file_read
